"""Search and filter utilities for CRM tables"""
import tkinter as tk
from tkinter import ttk
from typing import Callable, List, Dict, Any


class SearchFilter:
    """Search and filter widget for tables"""

    def __init__(self, parent, on_filter_change: Callable[[str], None]):
        """
        Initialize search filter widget

        Args:
            parent: Parent widget
            on_filter_change: Callback function when filter changes (receives search text)
        """
        self.on_filter_change = on_filter_change
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self._on_search_change)

        # Create frame
        self.frame = ttk.Frame(parent)

        # Search label and entry
        ttk.Label(self.frame, text="Search:").pack(side="left", padx=5)
        search_entry = ttk.Entry(self.frame, textvariable=self.search_var, width=40)
        search_entry.pack(side="left", padx=5)

        # Clear button
        ttk.Button(self.frame, text="Clear", command=self._clear_search).pack(side="left", padx=5)

    def _on_search_change(self, *args):
        """Handle search text change"""
        search_text = self.search_var.get()
        self.on_filter_change(search_text)

    def _clear_search(self):
        """Clear search text"""
        self.search_var.set("")

    def pack(self, **kwargs):
        """Pack the frame"""
        self.frame.pack(**kwargs)

    def get_search_text(self) -> str:
        """Get current search text"""
        return self.search_var.get()


class DataExporter:
    """Export table data to CSV and Excel formats"""

    @staticmethod
    def export_to_csv(filename: str, columns: List[str], rows: List[List[Any]]) -> bool:
        """
        Export data to CSV file

        Args:
            filename: Output filename
            columns: Column headers
            rows: List of row data

        Returns:
            True if successful, False otherwise
        """
        try:
            import csv
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(columns)
                writer.writerows(rows)
            return True
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False

    @staticmethod
    def export_to_excel(filename: str, columns: List[str], rows: List[List[Any]]) -> bool:
        """
        Export data to Excel file

        Args:
            filename: Output filename
            columns: Column headers
            rows: List of row data

        Returns:
            True if successful, False otherwise
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"

            # Add headers with styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col_num, header in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Add data rows
            for row_num, row_data in enumerate(rows, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)

            # Auto-adjust column widths
            for col_num, header in enumerate(columns, 1):
                max_length = len(str(header))
                for row in rows:
                    if col_num - 1 < len(row):
                        max_length = max(max_length, len(str(row[col_num - 1])))
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = min(max_length + 2, 50)

            wb.save(filename)
            return True
        except ImportError:
            print("openpyxl is not installed. Install it with: pip install openpyxl")
            return False
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False


def search_filter_rows(rows: List[Dict[str, Any]], search_text: str, search_fields: List[str]) -> List[Dict[str, Any]]:
    """
    Filter rows based on search text

    Args:
        rows: List of row dictionaries
        search_text: Search text (case-insensitive)
        search_fields: Fields to search in

    Returns:
        Filtered list of rows
    """
    if not search_text:
        return rows

    search_text = search_text.lower()
    filtered = []

    for row in rows:
        for field in search_fields:
            if field in row and search_text in str(row[field]).lower():
                filtered.append(row)
                break

    return filtered
